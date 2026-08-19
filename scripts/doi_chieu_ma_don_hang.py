"""
Đối chiếu Mã đơn hàng giữa nhãn giao hàng PDF (đã in thực tế) và file Excel
"đơn hàng giao" xuất từ Shopee — trả lời 2 câu hỏi:
  1. Có đơn nào đã in nhãn (PDF) nhưng KHÔNG có trong file Excel không?
     (thường do xuất file Excel quá sớm, trước khi đơn đó phát sinh)
  2. Có đơn nào có trong Excel nhưng CHƯA in nhãn (PDF) không?
     (thường là đơn đang "Chờ giao hàng" — không phải lỗi dữ liệu)

Đây là công cụ CHẠY TAY khi cần kiểm tra chéo — không phải 1 chức năng
trong web app, không tự động chạy định kỳ.

Cách dùng:
    python3 scripts/doi_chieu_ma_don_hang.py <thư_mục_chứa_PDF> <file_excel.xlsx>

Ví dụ:
    python3 scripts/doi_chieu_ma_don_hang.py "Misa hàng hóa/19.08" "Misa hàng hóa/19.08/file shopee 18.19-08.xlsx"

Yêu cầu: pip install pymupdf openpyxl (nếu chưa có)
"""

import sys
import re
import glob
import os
from collections import defaultdict

try:
    import pymupdf as fitz
except ImportError:
    import fitz  # bản pymupdf cũ đăng ký module dưới tên "fitz"
import openpyxl

# Cụm số dòng sản phẩm trong PDF, VD "1. Tên sản phẩm..., SL: 2" — neo vào
# đầu dòng (MULTILINE) để không bị dính vào số trong địa chỉ khách hàng
# (VD "Khóm 1 Số Nhà 44." từng gây lỗi nếu không neo đầu dòng).
ITEM_RE = re.compile(r"^(\d{1,2})\.\s(.+?),\s*SL:\s*(\d+)", re.DOTALL | re.MULTILINE)

# "Mã đơn hàng" là mã NGẮN dạng ngày+ký tự (VD "260819QUHNRTWD") — KHÁC với
# "Mã vận đơn" (VD "SPXVN060483136168" hoặc mã hãng vận chuyển khác). 2 nhãn
# PDF khác layout đã gặp thực tế:
#   - Nhãn thường: có cả "Mã vận đơn:" và "Mã đơn hàng:", 2 mã nằm cạnh nhau
#   - Nhãn Hỏa Tốc: chỉ có "Mã đơn hàng:", không có "Mã vận đơn:"
# Nên KHÔNG dò theo thứ tự chữ trong văn bản (dễ lẫn đúng-sai giữa 2 mã), mà
# dò theo toạ độ: mã nào cùng hàng (trục Y) với nhãn "Mã đơn hàng:" mới đúng.
ORDER_CODE_RE = re.compile(r"^\d{6}[A-Z0-9]{6,}$")
Y_TOLERANCE = 3.0


def extract_order_code(page) -> str | None:
    blocks = page.get_text("blocks")
    label_block = None
    for b in blocks:
        if "Mã đơn hàng:" in b[4]:
            label_block = b
            break
    if not label_block:
        return None
    label_y0, label_y1 = label_block[1], label_block[3]
    for b in blocks:
        text = b[4].strip()
        if not ORDER_CODE_RE.match(text):
            continue
        if abs(b[1] - label_y0) <= Y_TOLERANCE or abs(b[3] - label_y1) <= Y_TOLERANCE:
            return text
    return None


def extract_pdf_orders(pdf_folder: str) -> dict:
    """Trả về {mã đơn hàng: {"so_dong": N, "tong_sl": N, "file": ..., "trang": N}}"""
    orders = {}
    pdf_files = sorted(glob.glob(os.path.join(pdf_folder, "*.pdf")))
    for path in pdf_files:
        doc = fitz.open(path)
        for i in range(len(doc)):
            page = doc[i]
            code = extract_order_code(page)
            text = page.get_text()
            items = ITEM_RE.findall(text)
            tong_sl = sum(int(sl) for _, _, sl in items)
            loc = f"{os.path.basename(path)} trang {i + 1}"
            if not code:
                print(f"  !! Không dò được Mã đơn hàng ở {loc} — kiểm tra layout mới, dò tay dòng này.")
                continue
            if code in orders:
                print(f"  !! Mã đơn hàng {code} xuất hiện nhiều hơn 1 lần ({orders[code]['loc']} và {loc}) — có thể in trùng nhãn.")
                continue
            orders[code] = {"so_dong": len(items), "tong_sl": tong_sl, "loc": loc}
    return orders


def extract_excel_orders(excel_path: str) -> dict:
    """Trả về {mã đơn hàng: {"trang_thai": ..., "tong_sl": N, "so_dong": N}}"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    def col(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        raise ValueError(f'Không tìm thấy cột "{name}" trong file Excel — kiểm tra lại đúng file export đơn hàng từ Shopee')

    i_ma = col("Mã đơn hàng")
    i_trangthai = col("Trạng Thái Đơn Hàng")
    i_sl = col("Số lượng")

    orders = defaultdict(lambda: {"trang_thai": None, "tong_sl": 0, "so_dong": 0})
    for row in ws.iter_rows(min_row=2, values_only=True):
        ma = row[i_ma]
        if not ma:
            continue
        entry = orders[ma]
        entry["trang_thai"] = row[i_trangthai]
        entry["tong_sl"] += int(row[i_sl] or 0)
        entry["so_dong"] += 1
    return dict(orders)


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    pdf_folder, excel_path = sys.argv[1], sys.argv[2]

    print(f"Đang đọc PDF trong: {pdf_folder}")
    pdf_orders = extract_pdf_orders(pdf_folder)
    print(f"-> Trích được {len(pdf_orders)} Mã đơn hàng từ PDF.\n")

    print(f"Đang đọc Excel: {excel_path}")
    excel_orders = extract_excel_orders(excel_path)
    print(f"-> Có {len(excel_orders)} Mã đơn hàng trong Excel.\n")

    pdf_keys = set(pdf_orders.keys())
    excel_keys = set(excel_orders.keys())

    matched = pdf_keys & excel_keys
    pdf_only = pdf_keys - excel_keys
    excel_only = excel_keys - pdf_keys

    print("=" * 70)
    print(f"KHỚP CẢ 2 BÊN: {len(matched)} đơn")
    mismatch_sl = []
    for code in sorted(matched):
        p, e = pdf_orders[code], excel_orders[code]
        if p["tong_sl"] != e["tong_sl"]:
            mismatch_sl.append((code, p["tong_sl"], e["tong_sl"]))
    if mismatch_sl:
        print(f"  !! {len(mismatch_sl)} đơn khớp mã nhưng LỆCH tổng số lượng (PDF vs Excel):")
        for code, psl, esl in mismatch_sl:
            print(f"     {code}: PDF SL={psl}  Excel SL={esl}")
    else:
        print("  Tất cả các đơn khớp mã đều khớp luôn tổng số lượng.")

    print("\n" + "=" * 70)
    print(f"CÓ TRONG PDF NHƯNG KHÔNG CÓ TRONG EXCEL: {len(pdf_only)} đơn")
    print("  (Excel có thể đã xuất quá sớm, trước khi đơn này phát sinh — xuất lại file muộn hơn)")
    for code in sorted(pdf_only):
        print(f"   {code}  ({pdf_orders[code]['loc']}, tổng SL={pdf_orders[code]['tong_sl']})")

    print("\n" + "=" * 70)
    print(f"CÓ TRONG EXCEL NHƯNG KHÔNG CÓ TRONG PDF: {len(excel_only)} đơn")
    print("  (Bình thường nếu trạng thái là 'Chờ giao hàng' — đơn chưa in nhãn, chưa xuất kho)")
    for code in sorted(excel_only):
        e = excel_orders[code]
        flag = "" if e["trang_thai"] == "Chờ giao hàng" else "  <-- KHÁC 'Chờ giao hàng', kiểm tra lại"
        print(f"   {code}  (trạng thái: {e['trang_thai']}, tổng SL={e['tong_sl']}){flag}")

    print("\n" + "=" * 70)
    print(f"Tổng số lượng (Excel, các đơn khớp mã): {sum(excel_orders[c]['tong_sl'] for c in matched)}")
    print(f"Tổng số lượng (PDF, các đơn khớp mã):   {sum(pdf_orders[c]['tong_sl'] for c in matched)}")


if __name__ == "__main__":
    main()
